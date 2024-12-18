import openpyxl
# 导入 openpyxl 模块，用于操作 Excel 文件（如创建、读取、写入等）。
import os
# 导入 os 模块，用于文件和目录操作（例如检查文件是否存在）。
from openpyxl.styles import Alignment


# 从 openpyxl.styles 中导入 Alignment 类，用于设置单元格的对齐方式。

def parse_attendance_input(input_str: str, days_in_month: int = 31):
    # 定义函数 parse_attendance_input，用于解析考勤输入字符串。
    # 参数：input_str：输入字符串，例如 2-13,17-29,30.1,19+1。
    # days_in_month：一个月的天数，默认值为 31。
    """解析快速输入字符串"""
    attendance = {"morning": {}, "afternoon": {}, "overtime": {}}

    # 定义字典 attendance，包含三个子字典：
    # morning：存储上午的考勤数据。
    # afternoon：存储下午的考勤数据。
    # overtime：存储加班数据。

    def mark_attendance(day: int, period: int = None, overtime: float = None):
        """标记考勤数据"""
        # 定义嵌套函数 mark_attendance，用于在 attendance 字典中记录考勤信息。
        # 参数：
        # day：日期（如 1, 2, 3...）。
        # period：上午（1）或下午（2）；若为 None 表示全天。
        # overtime：加班时间，为浮点数；若为 None 表示无加班。

        if 1 <= day <= days_in_month:
            # 检查 day 是否在 1 到 days_in_month 范围内，超出范围的日期将被忽略。

            if period == 1:
                attendance["morning"][day] = "✓"
                # 如果 period 为 1（上午），在 attendance["morning"] 中将该日期的值标记为 "✓"

            elif period == 2:
                attendance["afternoon"][day] = "✓"
                # 如果 period 为 2（下午），在 attendance["afternoon"] 中将该日期的值标记为 "✓"

            else:  # 未指定上午或下午，默认全勤
                attendance["morning"][day] = "✓"
                attendance["afternoon"][day] = "✓"
                # 如果未指定 period，默认标记该日期的上午和下午均为 "✓"

            if overtime is not None:
                attendance["overtime"][day] = overtime
                # 如果提供了加班时间 overtime，将其记录到 attendance["overtime"]。

    for part in input_str.split(","):
        part = part.strip()
        if not part:
            continue
            # 遍历 input_str 中的每个片段：
            # 用 split(",") 按逗号分隔字符串，生成一个列表。
            # strip() 去除前后多余空格。
            # 跳过空字符串（if not part）。

        try:
            # 进入异常捕获块：用来处理格式错误的数据。

            if "+" in part and "." in part:  # 同时包含上午/下午和加班信息
                # 检查是否同时包含 +（加班信息）和 .（上午/下午标识）。
                # 例如：30.1+3。

                date_period, overtime = part.split("+")
                day, period = map(int, date_period.split("."))
                mark_attendance(day, period, float(overtime))
                # 将片段拆分为日期时间段和加班时间：
                # date_period：30.1
                # overtime：3
                # 再拆分 date_period 为 day 和 period。
                # 调用 mark_attendance 标记信息。

            elif "-" in part:  # 日期范围
                start, end = map(int, part.split("-"))
                for day in range(start, end + 1):
                    mark_attendance(day)
                    # 检查是否是日期范围（如 2-13）。
                    # 将范围拆分为起始日期 start 和结束日期 end。
                    # 遍历范围内的每一天，调用 mark_attendance 标记全天。

            elif "." in part:  # 指定上午或下午
                day, period = map(int, part.split("."))
                mark_attendance(day, period)
                # 检查是否包含上午/下午标识（如 30.1）。
                # 拆分为日期和时间段，并标记考勤。

            elif "+" in part:  # 处理加班信息
                day, overtime = part.split("+")
                mark_attendance(int(day), None, float(overtime))
                # 检查是否只包含加班信息（如 19+1）。
                # 拆分为日期和加班时间，并标记全天考勤及加班。

            else:  # 单一天数
                mark_attendance(int(part))
                # 默认处理单个日期（如 15）。
                # 直接标记全天。

        except ValueError:
            print(f"Invalid format: {part}")
            continue
        # 捕获格式错误，打印错误信息，跳过当前片段。

    # 补全未指定的日期为默认值
    for day in range(1, days_in_month + 1):
        attendance["morning"].setdefault(day, "✗")
        attendance["afternoon"].setdefault(day, "✗")
        attendance["overtime"].setdefault(day, "")
        # 补全未设置的日期：
        # morning 和 afternoon 的默认值为 "✗"（未打卡）。
        # overtime 的默认值为空字符串。

    return attendance
    # 返回完整的 attendance 数据。


def create_attendance_template(filename: str, days_in_month: int = 31):
    """创建考勤模板"""
    # 定义函数 create_attendance_template，用于创建 Excel 格式的考勤模板。
    # 参数：
    # filename：生成的文件名。
    # days_in_month：一个月的天数，默认值为 31。
    # """创建考勤模板"""：函数的文档字符串。

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤表"
    # 创建一个新的工作簿对象 wb：
    # openpyxl.Workbook()：生成一个新的 Excel 工作簿。
    # 获取默认的活动工作表 ws。
    # 将工作表的标题设置为 "考勤表"。

    # 填写表头
    ws.cell(row=1, column=1, value="员工编号")
    # 设置表格第 1 行第 1 列的单元格内容为 "员工编号"。

    for col in range(2, days_in_month + 2):
        ws.cell(row=1, column=col, value=col - 1)
    # 遍历从第 2 列到 days_in_month + 1 列的范围（总计 days_in_month 天）。
    # 每列代表一个日期：
    # col - 1：从 1 开始标记日期。

    set_column_width(ws, width=3.6, exclude_first_column=True)
    # 调用 set_column_width 函数：
    # 设置所有列宽为 3.6（除了第 1 列）。
    # 单元格内容居中。

    wb.save(filename)
    print(f"考勤模板已生成: {filename}")
    # 保存工作簿为指定文件名。
    # 打印提示信息，告知用户模板已生成。


def set_column_width(ws, width=3.6, exclude_first_column=False):
    """设置列宽，并将内容居中"""
    # 定义 set_column_width 函数，用于调整列宽并设置居中对齐。
    # 参数：
    # ws：工作表对象。
    # width：列宽，默认值为 3.6。
    # exclude_first_column：是否排除第 1 列，默认为 False。
    # """设置列宽，并将内容居中"""：函数的文档字符串。

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        if exclude_first_column and col_idx == 1:
            continue
        # 遍历工作表中的列：
        # 使用 enumerate 获取列的索引（从 1 开始）。
        # 如果 exclude_first_column 为 True 且当前列为第 1 列，则跳过。

        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = width
        # 获取当前列的字母（如 A、B、C）。
        # 使用 column_dimensions[col_letter].width 设置列宽。

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # 遍历工作表中的所有单元格，将内容居中对齐：
    # horizontal="center"：水平居中。
    # vertical="center"：垂直居中。


def fill_attendance(filename: str, employee_data: dict, days_in_month: int = 31):
    """根据员工考勤数据填写表格"""
    # 定义 fill_attendance 函数，用于根据员工考勤数据填写 Excel 表格。
    # 参数：
    # filename：Excel 文件名。
    # employee_data：员工考勤数据，字典格式，键为员工编号，值为考勤字符串。
    # days_in_month：一个月的天数，默认值为 31。
    # """根据员工考勤数据填写表格"""：函数的文档字符串。

    if not os.path.exists(filename):
        create_attendance_template(filename, days_in_month)
    # 检查文件是否存在：
    # 若不存在，调用 create_attendance_template 生成一个模板文件。

    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    # 加载指定的 Excel 文件。
    # 获取活动工作表。

    start_row = 2
    # 定义起始行：
    # 数据从第 2 行开始填写，第 1 行为表头。

    for emp_id, attendance_str in employee_data.items():
        # 填写员工编号
        ws.cell(row=start_row, column=1, value=emp_id)
        # 遍历员工考勤数据字典 employee_data。
        # 将员工编号填入第 1 列第 start_row 行。

        # 解析考勤数据
        attendance = parse_attendance_input(attendance_str, days_in_month)
        # 调用 parse_attendance_input 函数解析考勤字符串，返回解析后的数据。

        # 填写每个日期的考勤数据
        for col in range(2, days_in_month + 2):
            day = col - 1
            # 遍历每列，获取对应的日期（从 1 到 days_in_month）。

            # 上午
            ws.cell(row=start_row, column=col, value=attendance["morning"].get(day, "✗"))
            # 下午
            ws.cell(row=start_row + 1, column=col, value=attendance["afternoon"].get(day, "✗"))
            # 加班
            overtime = attendance["overtime"].get(day, "")
            if overtime:
                ws.cell(row=start_row + 2, column=col, value=overtime)

        start_row += 3  # 每个员工占3行

    # 设置列宽和居中（确保调整生效）
    set_column_width(ws, width=3.6, exclude_first_column=True)

    wb.save(filename)
    print(f"考勤表已更新并保存为: {filename}")


# 示例员工考勤数据
employees = {
    "11001": "2-13,17-29,30.1,19+1",
    "11002": "1-15,16.1+2.5,17-30",
    '11003': '1-18,19.2+2.5',
}

# 输出文件名
output_file = "考勤表.xlsx"

# 填充考勤数据
fill_attendance(output_file, employees)
