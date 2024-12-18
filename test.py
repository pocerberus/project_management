import openpyxl
import os
from openpyxl.styles import Alignment


def parse_attendance_input(input_str: str, days_in_month: int = 31):
    """解析快速输入字符串"""
    attendance = {"morning": {}, "afternoon": {}, "overtime": {}}

    def mark_attendance(day: int, period: int = None, overtime: float = None):
        """标记考勤数据"""
        if 1 <= day <= days_in_month:
            if period == 1:
                attendance["morning"][day] = "✓"
            elif period == 2:
                attendance["afternoon"][day] = "✓"
            else:  # 未指定上午或下午，默认全勤
                attendance["morning"][day] = "✓"
                attendance["afternoon"][day] = "✓"
            if overtime is not None:
                attendance["overtime"][day] = overtime

    for part in input_str.split(","):
        part = part.strip()
        if not part:
            continue

        try:
            if "+" in part and "." in part:  # 同时包含上午/下午和加班信息
                date_period, overtime = part.split("+")
                day, period = map(int, date_period.split("."))
                mark_attendance(day, period, float(overtime))
            elif "-" in part:  # 日期范围
                start, end = map(int, part.split("-"))
                for day in range(start, end + 1):
                    mark_attendance(day)
            elif "." in part:  # 指定上午或下午
                day, period = map(int, part.split("."))
                mark_attendance(day, period)
            elif "+" in part:  # 处理加班信息
                day, overtime = part.split("+")
                mark_attendance(int(day), None, float(overtime))
            else:  # 单一天数
                mark_attendance(int(part))
        except ValueError:
            print(f"Invalid format: {part}")
            continue

    # 补全未指定的日期为默认值
    for day in range(1, days_in_month + 1):
        attendance["morning"].setdefault(day, "✗")
        attendance["afternoon"].setdefault(day, "✗")
        attendance["overtime"].setdefault(day, "")

    return attendance


def create_attendance_template(filename: str, days_in_month: int = 31):
    """创建考勤模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤表"

    # 填写表头
    ws.cell(row=1, column=1, value="员工编号")
    for col in range(2, days_in_month + 2):
        ws.cell(row=1, column=col, value=col - 1)

    # 设置列宽（除第1列外，其他列宽度为 3.6）并居中
    set_column_width(ws, width=3.6, exclude_first_column=True)

    wb.save(filename)
    print(f"考勤模板已生成: {filename}")


def set_column_width(ws, width=3.6, exclude_first_column=False):
    """设置列宽，并将内容居中"""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        if exclude_first_column and col_idx == 1:
            continue
        col_letter = col_cells[0].column_letter
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


def fill_attendance(filename: str, employee_data: dict, days_in_month: int = 31):
    """根据员工考勤数据填写表格"""
    # 如果文件不存在，生成模板
    if not os.path.exists(filename):
        create_attendance_template(filename, days_in_month)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    start_row = 2  # 从第2行开始填写员工数据
    for emp_id, attendance_str in employee_data.items():
        # 填写员工编号
        ws.cell(row=start_row, column=1, value=emp_id)

        # 解析考勤数据
        attendance = parse_attendance_input(attendance_str, days_in_month)

        # 填写每个日期的考勤数据
        for col in range(2, days_in_month + 2):
            day = col - 1

            # 上午
            ws.cell(row=start_row, column=col, value=attendance["morning"].get(day, "✗"))
            # 下午
            ws.cell(row=start_row + 1, column=col, value=attendance["afternoon"].get(day, "✗"))
            # 加班
            overtime = attendance["overtime"].get(day, "")
            if overtime:
                ws.cell(row=start_row + 2, column=col, value=overtime)

        start_row += 3  # 每个员工占3行

    # 设置列宽和居中（确保调整生效）
    set_column_width(ws, width=3.6, exclude_first_column=True)

    wb.save(filename)
    print(f"考勤表已更新并保存为: {filename}")


# 示例员工考勤数据
employees = {
    "11001": "2-13,17-29,30.1,19+1",
    "11002": "1-15,16.1+2.5,17-30",
    '11003': '1-18,19.2+2.5',
}

# 输出文件名
output_file = "考勤表.xlsx"

# 填充考勤数据
fill_attendance(output_file, employees)
