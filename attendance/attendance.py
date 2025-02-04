import csv
from decimal import Decimal, InvalidOperation
import openpyxl  # 导入 openpyxl 模块，用于处理 Excel 文件
import os  # 导入 os 模块，用于操作文件和目录
import calendar  # 导入 calendar 模块，用于处理日期和时间
import logging  # 导入 logging 模块，用于记录日志信息
import mysql.connector
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.worksheet.page import PageMargins
from datetime import datetime  # 动态获取时间
from openpyxl.utils import get_column_letter
from collections import defaultdict


# 配置日志记录，设置日志级别为 INFO，格式为时间 - 日志级别 - 消息内容
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def sort_txt_by_number(input_file, attendance_file):
    try:
        # 检查文件是否存在
        if not os.path.exists(input_file):
            print(f"输入文件不存在：{input_file}")
            return

        with open(input_file, 'r', encoding='utf-8') as file:
            lines = [line.strip() for line in file if line.strip() and not line.startswith("#")]
            # for line in file
            # 遍历 file 文件对象的每一行，逐行读取内容。
            # file 是通过 open 打开的文件对象，支持迭代访问，每次返回一行字符串（包括换行符 \n）。
            #
            # line.strip()
            # strip() 是字符串的一个方法，用于移除字符串开头和结尾的所有空白字符（如空格、制表符 \t 和换行符 \n）。
            # 如果一行内容为 " example text \n"，调用 strip() 后会得到 "example text"。
            #
            # if line.strip()
            # 这是一个条件过滤器，表示只有在 line.strip() 的结果为真时（非空字符串），才会将该行加入列表。
            # 空行（如只有换行符或空格）经过 strip() 会变成空字符串 ""，在布尔上下文中为 False，因此会被过滤掉。
            #
            # [... for ... if ...]
            # 这是一个列表生成式，用于构建一个列表。
            # 它会将符合条件的 line.strip() 的结果依次加入到列表中

        sorted_lines = sorted(lines, key=lambda x: int(x.split(':')[0]))
        # 1.sorted(lines)
        # sorted是Python的内置函数，用于对可迭代对象进行排序。
        # 它会返回一个新的排序后的列表，不会修改原始列表lines。
        # 语法：sorted(iterable, key=None, reverse=False)
        #     iterable：要排序的对象，这里是lines列表。
        #     key：指定排序的规则（一个函数），这里通过lambda 表达式定义。
        #     reverse：默认False，表示升序排序。设置为True则降序
        #
        # 2.key = lambda x: int(x.split(':')[0])
        # key是一个函数，指定排序的依据。
        # 这里使用lambda x 定义了一个匿名函数，x是列表中每个元素的值。
        #
        # 3.拆解lambda x: int(x.split(':')[0])
        #     1 x.split(':')
        #
        #         split(':')是字符串的一个方法，用于将字符串按冒号: 分割成多个部分。
        #         返回值是一个列表，其中每部分是分割后的子字符串。
        #         示例："1:苹果".split(':') → ["1", "苹果"]
        #     2 x.split(':')[0]
        #         取分割后列表的第一个元素，即冒号前的部分。
        #         示例："1:苹果".split(':')[0] → "1"
        #     3 int(x.split(':')[0])
        #         将冒号前的部分从字符串转换为整数，用作排序的依据。
        #         示例：int("1") → 1
        #     4 lambda x 是 Python 中用于定义 匿名函数 的语法。它的功能类似于 def 关键字创建的普通函数，但更简洁。
        #         lambda 参数: 表达式
        #         lambda：关键字，用于声明一个匿名函数。
        #         参数：函数的输入，可以是一个或多个，多个参数用逗号分隔。
        #         表达式：函数的返回值，必须是一个单一表达式，不能包含复杂语句。
        #         匿名函数的特点
        #         没有名称：lambda 定义的函数是匿名的，常用在需要临时函数的场合。
        #         内联简洁：适合用在较简单的情况下，定义和使用往往写在同一行。
        #         自动返回：lambda 的表达式部分会自动作为返回值，无需使用 return。

        with open(attendance_file, 'w', encoding='utf-8') as file:
            file.write('\n'.join(sorted_lines))
        # 代码解析
        # 1.with open(attendance_file, 'w', encoding='utf-8') as file
        #     open(attendance_file, 'w', encoding='utf-8')
        #     打开（或创建）一个文件进行写操作。
        #     参数说明：
        #         attendance_file：指定的文件路径。
        #         'w'：表示写模式，会覆盖文件内容。如果文件不存在，会自动创建一个新文件。
        #         encoding = 'utf-8'：指定使用UTF - 8编码写入内容，保证对非ASCII字符（如中文）兼容。
        #     with 关键字
        #     with 是上下文管理器，负责在操作完成后自动关闭文件，无需手动调用 file.close()，即使在程序异常时也会安全关闭文件。
        # 2.file.write('\n'.join(sorted_lines))
        #     '\n'.join(sorted_lines)
        #         将sorted_lines列表中的元素拼接成一个字符串，每个元素之间用换行符 \n分隔。
        #         示例：
        #         sorted_lines = ["1:香蕉", "2:橙子", "3:苹果"]
        #         result = '\n'.join(sorted_lines)
        #         print(result)
        #         # 输出：
        #         # 1:香蕉
        #         # 2:橙子
        #         # 3:苹果
        #     file.write(...)
        #         将拼接后的字符串写入文件。
        #         如果sorted_lines是上例中的内容，文件的最终内容会是：
        #         1: 香蕉
        #         2: 橙子
        #         3: 苹果
        # 优点
        # 自动关闭文件：使用with 块，无论文件操作是否成功，都会自动释放资源。
        # 简洁高效：直接将排序后的列表转换为字符串并一次性写入文件。
        # 跨平台支持：指定编码为UTF - 8，兼容各种操作系统和语言环境。
        # 注意事项
        # 覆盖风险：如果attendance_file已存在，其内容会被覆盖。如果需要追加内容，可以将'w'替换为'a'（追加模式）。
        # 编码问题：如果内容中包含特殊字符，确保文件编码和操作系统的默认编码兼容。UTF-8是推荐选择。
        # 数据格式一致性：确保sorted_lines的内容已经按照所需的格式排序并去除了无效数据（如空行）。

        print(f"文件已成功排序并保存到 {attendance_file}")
    except Exception as e:
        print(f"发生错误：{e}")


def load_emp_attendance_from_txt(file_path):
    """从 TXT 文件加载员工考勤数据
    参数:
        file_path: 文件路径，指向包含员工数据的文本文件
    返回:
        一部字典，键为员工编号，值为对应的考勤字符串"""
    employees = {}  # 初始化一个空字典，用于存储员工编号和考勤数据
    if not os.path.exists(file_path):
        # 检查文件是否存在，如果不存在则记录错误信息并返回空字典
        logging.error(f"文件 {file_path} 不存在，请检查路径！")
        return employees
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            # 尝试以只读模式打开文件，编码格式为 UTF-8
            for line in file:
                line = line.strip()  # 去除每行的首尾空格
                if not line:
                    continue  # 跳过空行
                if ":" not in line:
                    # 如果行中没有冒号分隔符，记录警告信息并跳过
                    logging.warning(f"无效的员工数据格式: {line}")
                    continue
                try:
                    emp_id, attendance = line.split(':', 1)
                    # 按冒号分割为员工编号和考勤数据
                    # line.split(':') 是 Python 字符串的 split 方法，用于根据指定的分隔符（这里是 :）将字符串分割成一个列表。
                    # 第二个参数 1 是 maxsplit，表示最多分割 1 次，即字符串被拆分为最多 2 个部分。返回值是一个包含最多两个元素的列表
                    # emp_id, attendance 是解包赋值，将分割后的列表中第一个元素赋值给 emp_id，第二个元素赋值给 attendance
                    emp_id = emp_id.strip()  # 去除编号中的空格
                    attendance = attendance.strip()  # 去除考勤数据中的空格
                    if not emp_id.isalnum() or not attendance:
                        # 检查员工编号是否为字母数字组合，考勤数据是否为空
                        # emp_id.isalnum() 是一个字符串方法，用于检查字符串是否仅由字母和数字组成（即是否是“字母数字”字符串），且不能为空。
                        # not emp_id.isalnum() 表示字符串 emp_id 不是由纯字母和数字组成。如果 emp_id 包含空格、特殊字符（例如 @, # 等），或者为空字符串，条件为真
                        # not attendance检查变量 attendance 是否为假值。假值包括：""（空字符串）、None、False等。not attendance 表示 attendance 是空字符串或其他假值
                        logging.warning(f"无效的员工数据格式: {line}")
                        continue
                    employees[emp_id] = attendance  # 将解析出的员工编号和考勤数据加入字典
                    # 这行代码的作用是将 emp_id 作为键（Key），将 attendance 作为值（Value），存储到字典 employees 中
                    # employees：这是一个字典（dict）对象，用于存储键值对。
                    # emp_id：表示一个员工的唯一标识（如员工编号）。
                    # attendance：表示该员工的考勤信息（如“Present”或“Absent”）
                    # 这行代码的作用是将 emp_id 和 attendance 作为键值对加入到字典中。如果 emp_id 已经存在于字典中，则会更新其对应的值为新的 attendance
                except ValueError:
                    # 捕获分割过程中可能发生的错误并记录警告信息
                    logging.warning(f"无效的员工数据格式: {line}")
    except Exception as e:
        # 捕获文件读取过程中发生的任何异常并记录错误信息
        logging.error(f"加载员工数据时发生错误: {e}")

    return employees  # 返回包含员工数据的字典 employees[emp_id] = attendance


def parse_attendance_input(input_str: str, emp_ids: list, days_in_month: int):
    # 解析快速输入字符串，返回按员工编号组织的考勤数据字典

    # 初始化考勤数据
    attendance_data = {
        emp_id: {
            "morning": defaultdict(lambda: "\u2717"),  # 默认早班 ✗
            "afternoon": defaultdict(lambda: "\u2717"),  # 默认晚班 ✗
            "overtime": defaultdict(float)  # 默认无加班
        }
        for emp_id in emp_ids
    }

    def mark_attendance(emp_id, day: int, period: int = None, overtime: float = None):
        # 标记考勤数据
        if not (1 <= day <= days_in_month):
            return

        if period == 1:
            attendance_data[emp_id]["morning"][day] = "\u2713"  # 早班 ✓
        elif period == 2:
            attendance_data[emp_id]["afternoon"][day] = "\u2713"  # 晚班 ✓
        else:
            attendance_data[emp_id]["morning"][day] = "\u2713"
            attendance_data[emp_id]["afternoon"][day] = "\u2713"  # 全天 ✓

        if overtime is not None:
            attendance_data[emp_id]["overtime"][day] = int(overtime) if overtime.is_integer() else overtime  # 去掉小数点

    # 解析考勤输入
    for part in input_str.split(","):
        part = part.strip()
        if not part:
            continue

        try:
            overtime = None
            # **1. 解析 "(x.y-a.y)+overtime" 形式**
            if part.startswith("(") and ")+" in part:
                main_part, overtime_str = part.rstrip(")").split(")+", 1)
                main_part = main_part.lstrip("(")  # 去掉左括号
                try:
                    overtime = float(overtime_str)
                except ValueError:
                    logging.warning(f"无效加班时间格式: {overtime_str}")
                    overtime = None
            # **2. 解析 "x+overtime" 形式**
            elif "+" in part:
                main_part, overtime_str = part.split("+", 1)
                try:
                    overtime = float(overtime_str)
                except ValueError:
                    logging.warning(f"无效加班时间格式: {overtime_str}")
                    overtime = None
            else:
                main_part = part

            # 处理 (x.y - a.y) 结构
            if "-" in main_part and "." in main_part:
                start, end = main_part.split("-")
                start_day, start_period = map(int, start.split("."))
                end_day, end_period = map(int, end.split("."))

                if start_period != end_period:
                    logging.warning(f"起始和结束时段不匹配: {part}")
                    continue

                for day in range(start_day, end_day + 1):
                    for emp_id in emp_ids:
                        mark_attendance(emp_id, day, start_period, overtime)

            # 处理单个日期时段，如 "3.1"
            elif "." in main_part:
                day, period = map(int, main_part.split("."))
                for emp_id in emp_ids:
                    mark_attendance(emp_id, day, period, overtime)

            # 处理日期范围，如 "5-7"
            elif "-" in main_part:
                start, end = map(int, main_part.split("-"))
                for day in range(start, end + 1):
                    for emp_id in emp_ids:
                        mark_attendance(emp_id, day, None, overtime)

            # **新增：处理单个日期，如 "9+2"**
            elif main_part.isdigit():
                day = int(main_part)
                for emp_id in emp_ids:
                    mark_attendance(emp_id, day, None, overtime)

            else:
                logging.warning(f"无效格式: {part}")

        except ValueError:
            logging.warning(f"无效格式: {part}")

    return attendance_data


def parse_csv_config(file_path, default_value=None):
    """
    将只有两列的 CSV 文件解析为字典，允许第二列为空，并支持去除 BOM
    :param file_path: CSV 文件路径
    :param default_value: 第二列为空时的默认值（默认 None）
    :return: 字典，第一列为键，第二列为值（允许为空）
    """
    config_data = {}

    # 检查文件路径是否存在
    if not os.path.exists(file_path):
        logging.error(f"文件路径不存在: {file_path}")
        return config_data

    # 支持多种文件编码
    encodings = ["utf-8", "utf-8-sig", "windows-1254", "windows-1252", "GB18030"]
    for encoding in encodings:
        try:
            with open(file_path, mode="r", encoding=encoding, newline="") as csvfile:
                reader = csv.reader(csvfile)

                first_row = next(reader, None)  # 读取首行，避免空文件错误
                if first_row:
                    first_row[0] = first_row[0].lstrip("\ufeff")  # 去除 BOM

                    # 处理首行数据
                    key = first_row[0].strip()
                    value = first_row[1].strip() if len(first_row) > 1 and first_row[1].strip() else default_value
                    config_data[key] = value

                for row in reader:
                    if not row:  # 跳过空行
                        continue
                    key = row[0].strip()
                    value = row[1].strip() if len(row) > 1 and row[1].strip() else default_value
                    config_data[key] = value

            break  # 成功解析，跳出循环
        except (UnicodeDecodeError, IndexError) as e:
            logging.warning(f"使用编码 {encoding} 读取失败: {e}")
            continue
    else:
        logging.error(f"无法解析文件 {file_path}，尝试的编码均失败")

    return config_data


# 从config_data中获取指定月的天数 days_in_month,year,month
def get_days_in_month(config_data):
    # 从 config_data 字典获取年份和月份
    year = int(config_data['year'])  # 获取年份
    month = int(config_data['month'])  # 获取月份

    # 使用 calendar.monthrange 获取该月份的天数
    days_in_month: int
    _, days_in_month = calendar.monthrange(year, month)
    return year, month, days_in_month


def safe_decimal(value, default=Decimal("0")):
    """ 安全转换为 Decimal，避免解析错误 """
    try:
        return Decimal(str(value).strip()) if value else default
    except (InvalidOperation, ValueError):
        return default


def get_employee_info(config_data, emp_ids, csv_filename="emp_info.csv"):
    """ 根据员工工号列表获取员工信息，优先从 MySQL 读取，失败时从 CSV 读取 """

    if not emp_ids:
        logging.warning("员工ID列表为空，无法执行查询。")
        return {}

    employee_info = {}

    # **尝试从 MySQL 读取数据**
    conn = None
    try:
        if config_data:  # 仅在 config_data 存在时尝试连接数据库
            conn = mysql.connector.connect(
                host=config_data['host'],
                user=config_data['user'],
                password=config_data.get('password', ""),  # 确保密码为空时不报错
                database=config_data['database']
            )

            with conn.cursor(dictionary=True) as cursor:
                # 构建参数化查询
                placeholders = ", ".join(["%s"] * len(emp_ids))
                query = f"SELECT emp_id, job_type, name, unit_price FROM employees WHERE emp_id IN ({placeholders})"
                cursor.execute(query, emp_ids)
                employee_data = cursor.fetchall()

                if employee_data:
                    employee_info = {
                        str(emp['emp_id']): {
                            'job_type': emp['job_type'],
                            'name': emp['name'],
                            'unit_price': safe_decimal(emp['unit_price'])  # 统一转为 Decimal
                        }
                        for emp in employee_data
                    }
                    logging.info(f"从数据库获取的员工信息: {len(employee_info)} 条")
                else:
                    logging.warning("从数据库获取的数据为空，将尝试从 CSV 读取。")

    except mysql.connector.Error as db_error:
        logging.error(f"从数据库获取数据失败: {db_error}")
    finally:
        if conn:
            conn.close()
            logging.info("数据库连接已关闭。")

    # **如果数据库获取失败，则从 CSV 读取**
    if not employee_info and os.path.exists(csv_filename):
        logging.info(f"尝试从文件 {csv_filename} 加载数据...")
        try:
            # **支持多种编码格式**
            encodings = ["utf-8", "utf-8-sig", "windows-1254", "windows-1252", "GB18030"]
            for encoding in encodings:
                try:
                    with open(csv_filename, mode="r", encoding=encoding) as file:
                        csv_reader = csv.DictReader(file)

                        # **去除 UTF-8 BOM 并清理列名**
                        csv_reader.fieldnames = [name.lstrip("\ufeff").strip() for name in csv_reader.fieldnames]

                        required_columns = {"emp_id", "job_type", "name", "unit_price"}
                        if not required_columns.issubset(csv_reader.fieldnames):
                            raise ValueError(f"CSV 文件缺少必要列: {required_columns - set(csv_reader.fieldnames)}")

                        for row in csv_reader:
                            emp_id = str(row.get("emp_id")).strip()
                            if emp_id and emp_id in map(str, emp_ids):
                                employee_info[emp_id] = {
                                    "job_type": row.get("job_type", ""),
                                    "name": row.get("name", ""),
                                    "unit_price": safe_decimal(row.get("unit_price"))
                                }

                        logging.info(f"从 CSV ({encoding}) 读取到 {len(employee_info)} 条员工数据")
                        break  # 读取成功，退出编码尝试
                except (UnicodeDecodeError, ValueError) as e:
                    logging.warning(f"使用编码 {encoding} 读取 CSV 失败: {e}")
        except Exception as csv_error:
            logging.error(f"读取 CSV 文件失败: {csv_error}")

    elif not os.path.exists(csv_filename):
        logging.warning(f"CSV 文件 {csv_filename} 不存在！")
    print(employee_info)
    return employee_info

def set_column_widths(ws, days_in_month):  # 设置列宽
    # 设置第1列(A), 第2列(B), 第3列(C)的列宽
    ws.column_dimensions["A"].width = 3  # 第1列(A)宽度为3
    ws.column_dimensions["B"].width = 7  # 第2列(B)宽度为7
    ws.column_dimensions["C"].width = 13  # 第3列(C)宽度为13

    # 设置列宽为3.6
    for col in range(4, days_in_month + 5):
        col_letter = get_column_letter(col)  # 自动处理列号到字母的转换
        ws.column_dimensions[col_letter].width = 3.6

    # 设置列宽为6.8
    for col in range(days_in_month + 4, days_in_month + 7):
        col_letter = get_column_letter(col)  # 自动处理列号到字母的转换
        ws.column_dimensions[col_letter].width = 6.8

    # 设置合计工资列宽为10.5
    column_letter = get_column_letter(days_in_month + 8)
    ws.column_dimensions[column_letter].width = 10.5

    # 设置签名列宽为15
    column_letter = get_column_letter(days_in_month + 9)
    ws.column_dimensions[column_letter].width = 15


def create_attendance_template(attendance_data, config_data, days_in_month, filename: str):
    # 创建考勤模板
    # 参数:
    #     :param filename: 模板保存的文件路径
    #     :param days_in_month: 动态月天数int
    #     :param config_data: csv文件中解析而来的配置数据字典
    #     :param attendance_data: 解析txt而来的字典,用于设置模板框线

    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "考勤表"

    # 动态获取当前年月
    now = datetime.now()
    month = int(config_data.get("month", now.month))  # 默认为当前月
    year = int(config_data.get("year", now.year))  # 默认为当前年
    project = config_data.get("project", '')
    recorder = config_data.get("recorder", '')
    manager = config_data.get("manager", '')

    def get_merge_range(start_row, start_col, end_row, end_col):
        # 生成合并单元格的范围字符串
        # :param start_row: 起始行号（整数）
        # :param start_col: 起始列号（整数）
        # :param end_row: 结束行号（整数）
        # :param end_col: 结束列号（整数）
        # :return: 合并范围字符串（如 A5:C5）

        start_cell = f"{get_column_letter(start_col)}{start_row}"
        end_cell = f"{get_column_letter(end_col)}{end_row}"
        return f"{start_cell}:{end_cell}"

    # 合并第1行并填写标题
    merge_range = get_merge_range(1, 1, 1, days_in_month + 9)
    ws.merge_cells(merge_range)
    header_cell1 = ws.cell(row=1, column=1, value="考勤表")  # 设置合并单元格的内容为 "考勤表"
    header_cell1.font = Font(size=28, bold=True)  # 设置字体为 28 号
    ws.row_dimensions[1].height = 40  # 设置行高为 40
    header_cell1.alignment = Alignment(horizontal="center", vertical="center")

    # 合并第2行填写项目名称
    merge_range = get_merge_range(2, 1, 2, days_in_month + 6)
    ws.merge_cells(merge_range)
    header_cell2 = ws.cell(row=2, column=1)
    header_cell2.value = f"项目名称：{project}"
    header_cell2.alignment = Alignment(horizontal="left", vertical="center")

    # 合并第2行填写日期
    merge_range = get_merge_range(2, days_in_month + 7, 2, days_in_month + 9)
    ws.merge_cells(merge_range)
    header_cell3 = ws.cell(row=2, column=days_in_month + 7)
    header_cell3.value = f"{year}年{month}月"
    header_cell3.alignment = Alignment(horizontal="center", vertical="center")

    # 第三行填写表头数据
    headers = ['序\n号', '工号', ''] + [str(day) for day in range(1, days_in_month + 1)]
    headers += ["出勤\n天数", "加班\n计时", "合计\n工数", "人工\n单价", "合计\n工资", "签名"]
    for col, header in enumerate(headers, start=1):
        # enumerate 是 Python 中的一个内置函数，用于将一个可迭代对象（如列表、元组等）转换为一个枚举对象，返回一个由索引和元素组成的元组。
        # headers 是一个包含列标题的列表。
        # enumerate(headers, start=1) 表示从 1 开始为 headers 列表中的每个元素提供一个索引。col 表示列索引，header 表示 headers 列表中的当前元素（即列标题）
        cell = ws.cell(row=3, column=col, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # 设置单元格内容居中及自动换行
    ws.row_dimensions[3].height = 34  # 设置行高为 34

    # 第三行填写数据
    header_cell4 = ws.cell(row=3, column=3, value='       日期\n姓名')  # 获取合并单元格的起始单元格并设置内容
    header_cell4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # 设置内容居中及自动换行

    # 设置打印标题行（每页顶部打印该行）
    ws.print_title_rows = "1:3"  # 设置从第1行到第3行作为打印标题行

    # 调用设置列宽函数
    set_column_widths(ws, days_in_month)

    # 创建边框样式（四个方向的线）
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    # 设置要添加边框的单元格范围
    start_row = 1
    end_row = len(attendance_data) * 3 + 3
    start_column = 1
    end_column = days_in_month + 9
    # 给指定区域的所有单元格添加边框
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
        for cell in row:
            cell.border = thin_border

    # 姓名,日期单元格斜线
    cell = ws.cell(row=3, column=3)
    diagonal_border = Border(
        diagonal=Side(border_style="thin", color="000000"),  # 设置斜线样式
        diagonalDown=True,  # 从左上到右下的斜线
        diagonalUp=False  # 从左下到右上的斜线（这里未启用）
    )
    # 应用斜线样式到单元格
    cell.border = diagonal_border

    # 设置页边距
    ws.page_margins = PageMargins(
        left=0.5,  # 左边距（英寸）
        right=0.5,  # 右边距（英寸）
        top=0.5,  # 上边距（英寸）
        bottom=0.5,  # 下边距（英寸）
        header=0.3,  # 页眉到页面顶部的距离（英寸）
        footer=0.3  # 页脚到页面底部的距离（英寸）
    )

    # 设置页眉和页脚
    ws.oddHeader.left.text = ""
    ws.oddHeader.center.text = ""
    ws.oddHeader.right.text = ""

    ws.oddFooter.left.text = "备注：出勤：✔ 旷工：✘"
    ws.oddFooter.center.text = f"考勤员：{recorder}       项目经理：{manager}"
    ws.oddFooter.right.text = "第&P页，共&N页"

    # 设置打印方向为横向或纵向
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # 横向
    # ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 纵向

    # 设置打印为单页宽
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1  # 所有列适应一页宽
    ws.page_setup.fitToHeight = 0  # 行数不限页数

    wb.save(filename)  # 保存工作簿
    logging.info(f"考勤模板已生成: {filename}")  # 记录模板生成的信息


def fill_attendance(filename: str, attendance_data: dict, employee_info: dict, config_data: dict, days_in_month):
    # 填充考勤信息到员工数据中
    # 删除文件如果存在
    if os.path.exists(filename):
        os.remove(filename)
        print(f"文件已存在并被删除：{filename}")

    if not os.path.exists(filename):
        create_attendance_template(attendance_data, config_data, days_in_month, filename)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    start_row = 4
    start_value = 1  # 自增数字的起始值

    for emp_id, attendance in attendance_data.items():
        logging.info(f"正在处理员工 {emp_id} 的考勤数据...")

        # 写入序号并设置单元格格式
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 2, end_column=1)
        cell0 = ws.cell(row=start_row, column=1, value=start_value)
        cell0.alignment = Alignment(horizontal="center", vertical="center")
        # 更新序号
        start_value += 1

        # 写入员工工号并合并单元格数据
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + 2, end_column=2)
        cell1 = ws.cell(row=start_row, column=2, value=emp_id)  # 写入编号到左上角单元格
        cell1.alignment = Alignment(horizontal="center", vertical="center")  # 设置单元格内容居中

        # 填充每天的考勤数据
        total_morning, total_afternoon, total_overtime = 0, 0, 0
        for col in range(4, days_in_month + 4):
            day = col - 3
            morning = attendance["morning"].get(day, "\u2717")
            afternoon = attendance["afternoon"].get(day, "\u2717")
            overtime = attendance["overtime"].get(day, "")

            cell2 = ws.cell(row=start_row, column=col, value=morning)
            cell2.alignment = Alignment(horizontal="center", vertical="center")

            cell3 = ws.cell(row=start_row + 1, column=col, value=afternoon)
            cell3.alignment = Alignment(horizontal="center", vertical="center")

            if overtime:
                cell4 = ws.cell(row=start_row + 2, column=col, value=overtime)
                cell4.alignment = Alignment(horizontal="center", vertical="center")

            total_morning += 0.5 if morning == "\u2713" else 0
            total_afternoon += 0.5 if afternoon == "\u2713" else 0
            total_overtime += float(overtime or 0)

        # 计算额外统计信息
        attendance_days = total_morning + total_afternoon
        overtime_days = total_overtime / 6
        total_days = attendance_days + overtime_days

        emp_info = employee_info.get(emp_id, {})
        total_salary = total_days * float(emp_info.get('unit_price', 0) or 0)
        total_salary = round(total_salary, 2)  # 保留两位小数

        # 填充出勤天数,加班计时,合计工数,合计工资列的数据并设置格式
        ws.merge_cells(start_row=start_row, start_column=days_in_month + 4, end_row=start_row + 1,
                       end_column=days_in_month + 4)
        cell5 = ws.cell(row=start_row, column=days_in_month + 4, value=attendance_days)
        cell5.number_format = "0.00"  # 设置保留两位小数
        cell5.alignment = Alignment(horizontal="center", vertical="center")

        # 出勤天数下的加班换算天数
        ws.merge_cells(start_row=start_row + 2, start_column=days_in_month + 4, end_row=start_row + 2,
                       end_column=days_in_month + 4)
        cell6 = ws.cell(row=start_row + 2, column=days_in_month + 4, value=overtime_days)
        cell6.number_format = "0.00"  # 设置保留两位小数
        cell6.alignment = Alignment(horizontal="center", vertical="center")

        # 加班计时
        ws.merge_cells(start_row=start_row, start_column=days_in_month + 5, end_row=start_row + 2,
                       end_column=days_in_month + 5)
        cell7 = ws.cell(row=start_row, column=days_in_month + 5, value=total_overtime)
        cell7.number_format = "0.00"  # 设置保留两位小数
        cell7.alignment = Alignment(horizontal="center", vertical="center")

        # 合计工数
        ws.merge_cells(start_row=start_row, start_column=days_in_month + 6, end_row=start_row + 2,
                       end_column=days_in_month + 6)
        cell8 = ws.cell(row=start_row, column=days_in_month + 6, value=total_days)
        cell8.number_format = "0.00"  # 设置保留两位小数
        cell8.alignment = Alignment(horizontal="center", vertical="center")

        # 合计工资
        ws.merge_cells(start_row=start_row, start_column=days_in_month + 8, end_row=start_row + 2,
                       end_column=days_in_month + 8)
        cell9 = ws.cell(row=start_row, column=days_in_month + 8, value=total_salary)
        cell9.number_format = "0.00"  # 设置保留两位小数
        cell9.alignment = Alignment(horizontal="center", vertical="center")

        # 签名
        ws.merge_cells(start_row=start_row, start_column=days_in_month + 9, end_row=start_row + 2,
                       end_column=days_in_month + 9)
        cell10 = ws.cell(row=start_row, column=days_in_month + 9, value='')
        cell10.alignment = Alignment(horizontal="center", vertical="center")

        # 填充员工的基本信息（职位-姓名、人工单价）
        if emp_info:
            logging.info(f"正在填充员工信息: {emp_info}")
            # 职位-姓名
            ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row + 2, end_column=3)
            cell11 = ws.cell(row=start_row, column=3,
                             value=emp_info.get('job_type', '') + '-' + emp_info.get('name', ''))
            logging.info(f"填充职位-姓名: {emp_info.get('job_type', '')} {emp_info.get('name', '')}")
            cell11.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 人工单价
            ws.merge_cells(start_row=start_row, start_column=days_in_month + 7, end_row=start_row + 2,
                           end_column=days_in_month + 7)
            cell12 = ws.cell(row=start_row, column=days_in_month + 7, value=emp_info.get('unit_price', 0))
            logging.info(f"填充人工单价: {emp_info.get('unit_price', 0)}")
            cell12.number_format = "0.00"  # 设置保留两位小数
            cell12.alignment = Alignment(horizontal="center", vertical="center")

        start_row += 3  # 下一位员工的行

    wb.save(filename)
    logging.info(f"考勤表已更新并保存为: {filename}")


def parse_all_attendance(employees, days_in_month):  # 对parse_attendance_input函数的封装,如此调用简单,因为形参,不用考虑底层逻辑
    # 解析所有员工的考勤数据
    attendance_data = {}
    for emp_id, input_str in employees.items():
        emp_attendance = parse_attendance_input(input_str, [emp_id], days_in_month)
        attendance_data[emp_id] = emp_attendance[emp_id]
    return attendance_data


def main():

    # 文件路径
    # 自动获取当前脚本所在目录，并拼接文件路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(current_dir, 'input.txt')
    attendance_file = os.path.join(current_dir, 'emp_attendance.txt')
    sort_txt_by_number(input_file, attendance_file)

    base_dir = os.path.dirname(__file__)
    employee_file = os.path.join(base_dir, "emp_attendance.txt")
    config_file = os.path.join(base_dir, "config.csv")
    output_file = os.path.join(base_dir, "考勤表.xlsx")

    # 解析csv文件获得config_data字典,动态获取年份和月份
    config_data = parse_csv_config(config_file, default_value=None)
    year, month, days_in_month = get_days_in_month(config_data)

    # 加载员工数据和数据库配置
    employees = load_emp_attendance_from_txt(employee_file)

    # 获取员工信息
    emp_ids = list(employees.keys())
    employee_info = get_employee_info(config_data, emp_ids)

    # 解析考勤数据
    attendance_data = parse_all_attendance(employees, days_in_month)

    # 填充考勤表
    fill_attendance(output_file, attendance_data, employee_info, config_data, days_in_month)


if __name__ == "__main__":
    main()
