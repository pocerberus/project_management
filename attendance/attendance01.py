import openpyxl  # 导入 openpyxl 模块，用于处理 Excel 文件
import os  # 导入 os 模块，用于操作文件和目录
from openpyxl.styles import Alignment, Font  # 从 openpyxl.styles 导入 Alignment,Font 类，用于设置单元格对齐方式及字体
import calendar  # 导入 calendar 模块，用于处理日期和时间
import logging  # 导入 logging 模块，用于记录日志信息
import pymysql
import json
from openpyxl.utils import get_column_letter
import mysql.connector

# 配置日志记录，设置日志级别为 INFO，格式为时间 - 日志级别 - 消息内容
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


def sort_txt_by_number(input_file, output_file):
    try:
        # 检查文件是否存在
        if not os.path.exists(input_file):
            print(f"输入文件不存在：{input_file}")
            return

        with open(input_file, 'r', encoding='utf-8') as file:
            lines = [line.strip() for line in file if line.strip()]
            '''
            # for line in file
            # 遍历 file 文件对象的每一行，逐行读取内容。
            # file 是通过 open 打开的文件对象，支持迭代访问，每次返回一行字符串（包括换行符 \n）。

            # line.strip()
            # strip() 是字符串的一个方法，用于移除字符串开头和结尾的所有空白字符（如空格、制表符 \t 和换行符 \n）。
            # 如果一行内容为 " example text \n"，调用 strip() 后会得到 "example text"。

            # if line.strip()
            # 这是一个条件过滤器，表示只有在 line.strip() 的结果为真时（非空字符串），才会将该行加入列表。
            # 空行（如只有换行符或空格）经过 strip() 会变成空字符串 ""，在布尔上下文中为 False，因此会被过滤掉。

            # [... for ... if ...]
            # 这是一个列表生成式，用于构建一个列表。
            # 它会将符合条件的 line.strip() 的结果依次加入到列表中
            '''

        sorted_lines = sorted(lines, key=lambda x: int(x.split(':')[0]))
        '''
        # 1.sorted(lines)
        # sorted是Python的内置函数，用于对可迭代对象进行排序。
        # 它会返回一个新的排序后的列表，不会修改原始列表lines。
        # 语法：sorted(iterable, key=None, reverse=False)
        #     iterable：要排序的对象，这里是lines列表。
        #     key：指定排序的规则（一个函数），这里通过lambda 表达式定义。
        #     reverse：默认False，表示升序排序。设置为True则降序
        #
        # 2.key = lambda x: int(x.split(':')[0])
        # key是一个函数，指定排序的依据。
        # 这里使用lambda x 定义了一个匿名函数，x是列表中每个元素的值。
        #
        # 3.拆解lambda x: int(x.split(':')[0])
        #     1 x.split(':')
        #
        #         split(':')是字符串的一个方法，用于将字符串按冒号: 分割成多个部分。
        #         返回值是一个列表，其中每部分是分割后的子字符串。
        #         示例："1:苹果".split(':') → ["1", "苹果"]
        #     2 x.split(':')[0]
        #         取分割后列表的第一个元素，即冒号前的部分。
        #         示例："1:苹果".split(':')[0] → "1"
        #     3 int(x.split(':')[0])
        #         将冒号前的部分从字符串转换为整数，用作排序的依据。
        #         示例：int("1") → 1
        #     4 lambda x 是 Python 中用于定义 匿名函数 的语法。它的功能类似于 def 关键字创建的普通函数，但更简洁。
        #         lambda 参数: 表达式
        #         lambda：关键字，用于声明一个匿名函数。
        #         参数：函数的输入，可以是一个或多个，多个参数用逗号分隔。
        #         表达式：函数的返回值，必须是一个单一表达式，不能包含复杂语句。
        #         匿名函数的特点
        #         没有名称：lambda 定义的函数是匿名的，常用在需要临时函数的场合。
        #         内联简洁：适合用在较简单的情况下，定义和使用往往写在同一行。
        #         自动返回：lambda 的表达式部分会自动作为返回值，无需使用 return。
        '''

        with open(output_file, 'w', encoding='utf-8') as file:
            file.write('\n'.join(sorted_lines))
        '''
        # 代码解析
        # 1.with open(output_file, 'w', encoding='utf-8') as file
        #     open(output_file, 'w', encoding='utf-8')
        #     打开（或创建）一个文件进行写操作。
        #     参数说明：
        #         output_file：指定的文件路径。
        #         'w'：表示写模式，会覆盖文件内容。如果文件不存在，会自动创建一个新文件。
        #         encoding = 'utf-8'：指定使用UTF - 8编码写入内容，保证对非ASCII字符（如中文）兼容。
        #     with 关键字
        #     with 是上下文管理器，负责在操作完成后自动关闭文件，无需手动调用 file.close()，即使在程序异常时也会安全关闭文件。
        # 2.file.write('\n'.join(sorted_lines))
        #     '\n'.join(sorted_lines)
        #         将sorted_lines列表中的元素拼接成一个字符串，每个元素之间用换行符 \n分隔。
        #         示例：
        #         sorted_lines = ["1:香蕉", "2:橙子", "3:苹果"]
        #         result = '\n'.join(sorted_lines)
        #         print(result)
        #         # 输出：
        #         # 1:香蕉
        #         # 2:橙子
        #         # 3:苹果
        #     file.write(...)
        #         将拼接后的字符串写入文件。
        #         如果sorted_lines是上例中的内容，文件的最终内容会是：
        #         1: 香蕉
        #         2: 橙子
        #         3: 苹果
        # 优点
        # 自动关闭文件：使用with 块，无论文件操作是否成功，都会自动释放资源。
        # 简洁高效：直接将排序后的列表转换为字符串并一次性写入文件。
        # 跨平台支持：指定编码为UTF - 8，兼容各种操作系统和语言环境。
        # 注意事项
        # 覆盖风险：如果output_file已存在，其内容会被覆盖。如果需要追加内容，可以将'w'替换为'a'（追加模式）。
        # 编码问题：如果内容中包含特殊字符，确保文件编码和操作系统的默认编码兼容。UTF-8是推荐选择。
        # 数据格式一致性：确保sorted_lines的内容已经按照所需的格式排序并去除了无效数据（如空行）。
        '''
        print(f"文件已成功排序并保存到 {output_file}")
    except Exception as e:
        print(f"发生错误：{e}")


# 自动获取当前脚本所在目录，并拼接文件路径
current_dir = os.path.dirname(os.path.abspath(__file__))
input_file = os.path.join(current_dir, 'input.txt')
output_file = os.path.join(current_dir, 'employees.txt')
sort_txt_by_number(input_file, output_file)
'''
    # 逐行解析
    # 1. current_dir = os.path.dirname(os.path.abspath(__file__))
    # os.path.abspath(__file__)：
    # 获取当前脚本文件的绝对路径（包含文件名）。
    # 例如，假设脚本位于 /home/user/project/script.py，则 os.path.abspath(__file__) 的结果是 /home/user/project/script.py。
    # os.path.dirname()：
    # 获取文件所在的目录路径。
    # 结合上例，os.path.dirname("/home/user/project/script.py") 的结果是 /home/user/project。
    # current_dir：
    # 最终保存了当前脚本所在的目录路径。

    # 2. input_file = os.path.join(current_dir, 'input.txt')
    # os.path.join()：
    # 将目录路径 current_dir 与文件名 input.txt 拼接成一个完整的路径。
    # 例如，如果 current_dir 是 /home/user/project，则结果是 /home/user/project/input.txt。
    # input_file：
    # 保存了待处理文件 input.txt 的完整路径。

    # 3. output_file = os.path.join(current_dir, 'employees.txt')
    # 功能与上面类似，只不过目标文件是 sorted_output.txt。
    # output_file 保存了排序后输出文件的完整路径。

    # 4. sort_txt_by_number(input_file, output_file)
    # 调用定义好的函数 sort_txt_by_number，并传入两个参数：
    # input_file：表示要读取并排序的输入文件路径。
    # output_file：表示排序后结果要保存的输出文件路径。
'''


def load_employees_from_txt(file_path):
    """从 TXT 文件加载员工考勤数据
    参数:
        file_path: 文件路径，指向包含员工数据的文本文件
    返回:
        一部字典，键为员工编号，值为对应的考勤字符串"""
    employees = {}  # 初始化一个空字典，用于存储员工编号和考勤数据
    if not os.path.exists(file_path):
        # 检查文件是否存在，如果不存在则记录错误信息并返回空字典
        logging.error(f"文件 {file_path} 不存在，请检查路径！")
        return employees
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            # 尝试以只读模式打开文件，编码格式为 UTF-8
            for line in file:
                line = line.strip()  # 去除每行的首尾空格
                if not line:
                    continue  # 跳过空行
                if ":" not in line:
                    # 如果行中没有冒号分隔符，记录警告信息并跳过
                    logging.warning(f"无效的员工数据格式: {line}")
                    continue
                try:
                    emp_id, attendance = line.split(':', 1)
                    # 按冒号分割为员工编号和考勤数据
                    # line.split(':') 是 Python 字符串的 split 方法，用于根据指定的分隔符（这里是 :）将字符串分割成一个列表。
                    # 第二个参数 1 是 maxsplit，表示最多分割 1 次，即字符串被拆分为最多 2 个部分。返回值是一个包含最多两个元素的列表
                    # emp_id, attendance 是解包赋值，将分割后的列表中第一个元素赋值给 emp_id，第二个元素赋值给 attendance
                    emp_id = emp_id.strip()  # 去除编号中的空格
                    attendance = attendance.strip()  # 去除考勤数据中的空格
                    if not emp_id.isalnum() or not attendance:
                        # 检查员工编号是否为字母数字组合，考勤数据是否为空
                        # emp_id.isalnum() 是一个字符串方法，用于检查字符串是否仅由字母和数字组成（即是否是“字母数字”字符串），且不能为空。
                        # not emp_id.isalnum() 表示字符串 emp_id 不是由纯字母和数字组成。如果 emp_id 包含空格、特殊字符（例如 @, # 等），或者为空字符串，条件为真
                        # not attendance检查变量 attendance 是否为假值。假值包括：""（空字符串）、None、False等。not attendance 表示 attendance 是空字符串或其他假值
                        logging.warning(f"无效的员工数据格式: {line}")
                        continue
                    employees[emp_id] = attendance  # 将解析出的员工编号和考勤数据加入字典
                    # 这行代码的作用是将 emp_id 作为键（Key），将 attendance 作为值（Value），存储到字典 employees 中
                    # employees：这是一个字典（dict）对象，用于存储键值对。
                    # emp_id：表示一个员工的唯一标识（如员工编号）。
                    # attendance：表示该员工的考勤信息（如“Present”或“Absent”）
                    # 这行代码的作用是将 emp_id 和 attendance 作为键值对加入到字典中。如果 emp_id 已经存在于字典中，则会更新其对应的值为新的 attendance
                except ValueError:
                    # 捕获分割过程中可能发生的错误并记录警告信息
                    logging.warning(f"无效的员工数据格式: {line}")
    except Exception as e:
        # 捕获文件读取过程中发生的任何异常并记录错误信息
        logging.error(f"加载员工数据时发生错误: {e}")
    return employees  # 返回包含员工数据的字典 employees[emp_id] = attendance


def parse_attendance_input(input_str: str, days_in_month: int = 31):
    """解析快速输入字符串
    参数:
        input_str: 表示考勤信息的字符串，例如 "1, 2.1, 3.2+2, 4-6"
        days_in_month: 当前月份的天数，默认为 31 天
    返回:
        一个包含早班、晚班和加班信息的嵌套字典"""
    attendance = {"morning": {}, "afternoon": {}, "overtime": {}}  # 初始化嵌套字典

    def mark_attendance(day: int, period: int = None, overtime: float = None):
        """标记考勤数据
        参数:
            day: 天数，对应日期
            period: 考勤时间段，1 表示早班，2 表示晚班，默认标记全天
            overtime: 加班时长，默认为 None
            这个函数只处理上午、下午、加班的情况，其他情况在下面外部逻辑中处理
        """
        if 1 <= day <= days_in_month:
            if period == 1:
                attendance["morning"][day] = "\u2713"  # 早班出勤标记
            elif period == 2:
                attendance["afternoon"][day] = "\u2713"  # 晚班出勤标记
            else:
                attendance["morning"][day] = "\u2713"
                attendance["afternoon"][day] = "\u2713"  # 全天出勤

            if overtime is not None:
                attendance["overtime"][day] = overtime  # 记录加班时间

    # 本段代码属于外部逻辑
    for part in input_str.split(","):
        part = part.strip()  # 去除每部分的首尾空格
        if not part:
            continue  # 跳过空部分
        try:
            if "+" in part and "." in part:
                # 如果部分同时包含 "+" 和 "."，表示 "日.时段+加班"
                date_period, overtime = part.split("+")
                day, period = map(int, date_period.split("."))
                # map(int, ...) 用来将 split() 返回的每个字符串转换成整数(相当于去掉了数字上的引号变成能计算的数字)
                mark_attendance(day, period, float(overtime))
                # 传递参数给mark_attendance函数
            elif "-" in part:
                # 如果部分包含 "-"，表示 "起始日-结束日"
                start, end = map(int, part.split("-"))
                for day in range(start, end + 1):
                    mark_attendance(day)  # 传递参数给mark_attendance函数
            elif "." in part:
                # 如果部分包含 "."，表示 "日.时段"
                day, period = map(int, part.split("."))
                mark_attendance(day, period)  # 传递参数给mark_attendance函数
            elif "+" in part:
                # 如果部分包含 "+"，表示 "日+加班"
                day, overtime = part.split("+")
                mark_attendance(int(day), None, float(overtime))  # 中间是period，但是没有值，所以写none
            else:
                # 否则部分表示单独的日期
                mark_attendance(int(part))  # 传递参数给mark_attendance函数
        except ValueError:
            logging.warning(
                f"无效的格式: {part}。预期格式包括 'day', 'day.period', 'day+overtime', 'start-end' 等。")
            continue

    # 此段代码为外部逻辑
    # 填充默认值，将未标记的日期标记为未出勤
    for day in range(1, days_in_month + 1):
        attendance["morning"].setdefault(day, "\u2717")
        attendance["afternoon"].setdefault(day, "\u2717")
        attendance["overtime"].setdefault(day, "")

    return attendance  # 返回解析后的考勤数据字典


def load_db_config(config_file):
    """从 JSON 配置文件加载 MySQL 数据库连接信息"""
    try:
        with open(config_file, 'r', encoding='utf-8') as file:
            config = json.load(file)
            # 这里json.load(file): 读取文件内容并解析为 Python 字典,这是这个函数的功能,想要使用这部字典后续需要定义变量让这个函数赋值
        return config
    except Exception as e:
        print(f"读取配置文件时出错：{e}")
        return None


def get_employee_info_from_mysql(config, emp_ids):
    """根据员工工号列表从 MySQL 获取员工信息，得到一部键为员工工号的字典"""
    if not config:
        print("数据库配置无效，无法连接数据库。")
        return {}

    if not emp_ids:
        print("员工ID列表为空，无法执行查询。")
        return {}

    try:
        # 连接数据库
        conn = mysql.connector.connect(
            host=config['host'],
            user=config['user'],
            password=config['password'],
            database=config['database']
        )

        # 自动管理游标和连接
        with conn.cursor(dictionary=True) as cursor:
            # 创建游标对象 cursor
            # 参数 dictionary=True 指定查询结果以字典形式返回，而非默认的元组形式
            # 使用 with 语句确保游标在代码块结束后自动关闭，避免资源泄露

            # 构建参数化查询
            placeholders = ", ".join(["%s"] * len(emp_ids))
            # 根据 emp_ids 的长度，动态生成 SQL 占位符字符串。
            # 假如 emp_ids = [101, 102]，生成的字符串为 "%s, %s"
            # 使用占位符 %s 实现参数化查询，避免直接拼接字符串，防止 SQL 注入风险

            query = f"SELECT emp_id, job_type, name, unit_price FROM employees WHERE emp_id IN ({placeholders})"
            # 构建查询语句，将占位符插入 WHERE emp_id IN (...) 中,保证查询语句安全且易于扩展保证查询语句安全且易于扩展

            # 执行查询
            cursor.execute(query, emp_ids)
            # query, emp_ids都为参数
            # 执行 SQL 查询，emp_ids 中的值会替换占位符 %s,通过参数化查询方式，确保数据被安全转义，防止 SQL 注入

            # 获取查询结果
            employee_data = cursor.fetchall()
            # 使用 fetchall() 获取查询结果，返回一个包含所有记录的列表

        # 关闭数据库连接
        conn.close()

        # 将结果转换为字典形式
        employee_info = {
            emp['emp_id']: {'job_type': emp['job_type'], 'name': emp['name'], 'unit_price': emp['unit_price']}
            for emp in employee_data
        }
        # emp 是字典推导式中的临时变量，不需要在上下文中单独定义。
        # 它的值来自 for emp in employee_data，每次迭代时会被动态赋值为 employee_data 的当前元素。
        # 列表推导式的基本语法：[expression for item in iterable if condition]
        # 它的作用域仅限于字典推导式或循环语句中，执行完成后就会释放,
        return employee_info

    except mysql.connector.Error as err:
        print(f"数据库连接错误: {err}")
        return {}


def main():
    # 从 JSON 配置文件加载数据库连接信息
    db_config_file = 'db_config.json'  # 假设配置文件路径
    db_config = load_db_config(db_config_file)

    if not db_config:
        return

    # 给出txt文件路径，调用函数加载员工工号
    file_path = 'employees.txt'  # TXT 文件路径
    emp_ids = list(employees.keys())
    # load_employees_from_txt 函数中employees字典已经return，外部可以直接访问这个字典

    if not emp_ids:
        print("未找到任何员工工号")
        return

    # 根据工号从 MySQL 获取员工信息
    employee_info = get_employee_info_from_mysql(db_config, emp_ids)

    # 输出员工信息
    for emp_id, info in employee_info.items():
        print(f"工号: {emp_id}, 工种: {info['job_type']}, 姓名: {info['name']}, 单价: {info['unit_price']}")


def create_attendance_template(filename: str, days_in_month: int = 31):
    """创建考勤模板
    参数:
        filename: 模板保存的文件路径
        days_in_month: 当前月份的天数"""
    wb = openpyxl.Workbook()  # 创建一个新的工作簿
    ws = wb.active  # 获取活动工作表
    ws.title = "考勤表"  # 设置表格标题

    ws.cell(row=1, column=1, value="工号")  # 填写标题行的第一列
    for col in range(2, days_in_month + 2):
        ws.cell(row=1, column=col, value=col - 1)  # 按日期填写列标题
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 首行标题进行居中处理

    set_column_width(ws, width=3.6, exclude_first_column=True)  # 设置列宽
    wb.save(filename)  # 保存工作簿
    logging.info(f"考勤模板已生成: {filename}")  # 记录模板生成的信息


def set_column_width(ws, width=3.6, exclude_first_column=False):
    """设置列宽，并将内容居中
    参数:
        ws: 工作表对象
        width: 列宽的数值
        exclude_first_column: 是否排除第一列"""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        if exclude_first_column and col_idx == 1:
            continue
        col_letter = col_cells[0].column_letter  # 获取列字母
        ws.column_dimensions[col_letter].width = width  # 设置列宽

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")  # 居中对齐


def fill_attendance(filename: str, employee_data: dict, days_in_month: int = 31):
    """根据员工考勤数据填写表格，并增加出勤统计列
    参数:
        filename: 表格文件路径
        employee_data: 包含员工编号及考勤数据的字典
        days_in_month: 当前月份的天数"""
    if not os.path.exists(filename):
        # 如果文件不存在，先创建考勤模板
        create_attendance_template(filename, days_in_month)

    wb = openpyxl.load_workbook(filename)  # 打开现有的 Excel 工作簿
    ws = wb.active  # 获取活动工作表

    extra_columns = [("出勤天数", 0), ("加班计时", 1), ("合计工数", 2)]
    base_col = days_in_month + 2  # 额外列的起始列号
    for title, offset in extra_columns:
        col_idx = base_col + offset
        ws.cell(row=1, column=col_idx, value=title)  # 增加新地统计列标题

    start_row = 2  # 设置起始行，从第二行开始填写数据

    for emp_id, attendance_str in employee_data.items():
        ws.cell(row=start_row, column=1, value=emp_id)  # 填写员工编号到第一列

        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + 2, end_column=1)
        cell = ws.cell(row=start_row, column=1)
        cell.alignment = Alignment(horizontal="center", vertical="center")  # 将编号单元格内容居中对齐

        attendance = parse_attendance_input(attendance_str, days_in_month)  # 解析员工的考勤数据字符串

        total_morning = 0
        total_afternoon = 0
        total_overtime = 0
        for col in range(2, days_in_month + 2):
            day = col - 1  # 列号减去 1 对应日期
            morning = attendance["morning"].get(day, "\u2717")
            ws.cell(row=start_row, column=col, value=morning)  # 填写早班考勤
            if morning == "\u2713":
                total_morning += 0.5

            afternoon = attendance["afternoon"].get(day, "\u2717")
            ws.cell(row=start_row + 1, column=col, value=afternoon)  # 填写晚班考勤
            if afternoon == "\u2713":
                total_afternoon += 0.5

            overtime = attendance["overtime"].get(day, "")
            if overtime:
                ws.cell(row=start_row + 2, column=col, value=overtime)  # 填写加班数据
                total_overtime += float(overtime)

        attendance_days = total_morning + total_afternoon
        overtime_days = total_overtime / 6
        total_days = attendance_days + overtime_days

        ws.merge_cells(start_row=start_row, start_column=base_col, end_row=start_row + 1, end_column=base_col)
        attendance_cell = ws.cell(row=start_row, column=base_col, value=attendance_days)
        attendance_cell.alignment = Alignment(horizontal="center", vertical="center")  # 填写出勤天数（前两行合并，显示总出勤天数）

        overtime_days_cell = ws.cell(row=start_row + 2, column=base_col, value=overtime_days)
        overtime_days_cell.alignment = Alignment(horizontal="center", vertical="center")  # 填写加班天数（出勤天数第3行显示加班时长换算的出勤天数）
        ws.cell(row=start_row + 2, column=base_col, value=overtime_days).number_format = "0.00"

        ws.merge_cells(start_row=start_row, start_column=base_col + 1, end_row=start_row + 2, end_column=base_col + 1)
        overtime_cell = ws.cell(row=start_row, column=base_col + 1, value=total_overtime)
        overtime_cell.alignment = Alignment(horizontal="center", vertical="center")  # 填写加班计时（所有加班小时数合并）

        ws.merge_cells(start_row=start_row, start_column=base_col + 2, end_row=start_row + 2, end_column=base_col + 2)
        total_cell = ws.cell(row=start_row, column=base_col + 2, value=total_days)
        total_cell.alignment = Alignment(horizontal="center", vertical="center")  # 填写合计工数（出勤天数 + 加班天数，合并）

        start_row += 3  # 每个员工占用三行，因此起始行向下移动三行

    wb.save(filename)  # 保存更新后的工作簿
    logging.info(f"考勤表已更新并保存为: {filename}")  # 记录考勤表更新的信息


if __name__ == "__main__":
    year, month = 2024, 12  # 动态获取年份和月份对应的天数
    days_in_month = calendar.monthrange(year, month)[1]  # 获取月份的天数

    employee_file = os.path.join(os.path.dirname(__file__), "employees.txt")  # 加载员工数据文件
    employees = load_employees_from_txt(employee_file)  # 从文件中加载员工考勤数据

    output_file = os.path.join(os.path.dirname(__file__), "考勤表.xlsx")  # 输出文件名

    fill_attendance(output_file, employees, days_in_month)  # 填充考勤数据
